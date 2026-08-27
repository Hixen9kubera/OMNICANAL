"""
packing_indice.py — Índice de UN packing list, listo para buscarle el renglón a
un SKU.

Es la mitad "archivo" del validador de costos de PUBLICADOS EN MERCADO LIBRE
(v0.281.0), que corre el Resolver al revés: en vez de cargar un xlsx y empatar
sus renglones contra el catálogo, se parte de un SKU y se le busca su renglón.
Para eso el archivo tiene que estar indexado por tres cosas a la vez:

  · **fotos** con su sha256 y su dHash — el empate de verdad, porque la foto de
    Odoo y la del packing list son el MISMO archivo en el 92% de los casos;
  · **texto crudo** por renglón — lo que se le da a leer al modelo;
  · **números POR CAJA** — de donde sale el costo.

Los renglones y las fotos los pone :mod:`packing_parser` tal cual (la
asociación foto→renglón por anclas del ZIP ya está resuelta ahí y no se
reimplementa). Lo que este módulo agrega es la segunda lectura del libro con su
propio juego de patrones de columna, y ahí está el porqué de todo el archivo:

    Con OOCU8248653, `mapear_columnas` no reconocía 件数 / 产品申报数量 /
    产品总件数 / 总重量, y cajas, piezas, peso y CBM salían en CERO. Un flete
    en cero no avisa — solo subestima el costo, y un costo bajo se convierte en
    un precio bajo.

TRES CORRECCIONES sobre el script que corrió esto a mano, y las tres importan:

1. **El encabezado se LOCALIZA** con ``packing_parser.encontrar_encabezado()``.
   El script leía ``ws[1]`` fijo; con un archivo que trae membrete arriba —que
   son la mayoría— todas las columnas numéricas salen None y el costo se va al
   suelo EN SILENCIO. Es el defecto más caro de portar.
2. **La hoja es ``wb.active``**, no ``wb.sheetnames[0]``: es la que lee el
   parser, y si no coinciden se estarían indexando dos hojas distintas.
3. **Los grupos de caja compartida se calculan UNA vez.** El script recorría
   ``ws.merged_cells.ranges`` por cada renglón: cuadrático, y estos archivos
   traen miles de renglones con cientos de merges.

LA FÓRMULA — está toda en :meth:`Indice.datos` y no se negocia. Todo se empareja
POR CAJA: ``piezas_en_caja = total_fila / num_cajas``, ``piezas_grupo`` es la
suma de las piezas del grupo que comparte cartón, y de ahí salen
``cbm_por_pieza`` y ``peso_pieza``. Mezclar el CBM TOTAL del renglón con las
piezas de UNA caja multiplica el flete por el número de cajas: ese error dio
$187.69 donde lo real eran $79.02.

Por eso mismo el ``cbm_por_pieza`` que devuelve ``packing_parser.leer()`` NO se
usa aquí: viene multiplicado por las cajas cuando la columna de piezas es "por
caja", y ni ``normalizar_semantica`` ni ``completar`` lo corrigen después.
"""
from __future__ import annotations

import hashlib
import io
import logging
import re
from typing import Any

import openpyxl

from services import packing_costos, packing_parser

log = logging.getLogger("omnicanal.packing.indice")

# ── Patrones de columna ──────────────────────────────────────────────────────
# Se suman a los de packing_parser.mapear_columnas, que se quedan como están
# para no mover al Resolver clásico bajo los pies. Aquí se resuelven a número de
# columna sobre el encabezado real, no sobre la fila 1.
# "gross weight" / "g.w." entran AQUÍ, en el peso de la FILA, no en el de la
# caja: es como lo escriben los packing lists en inglés (PCIU9532241 pone
# `Gross weight`). Ponerlo en PAT_CAJA sería el error caro —un 毛重 suelto casa
# con 总毛重 y da pesos 10x—, y aquí no hace daño porque `c_caja`, que exige la
# marca explícita de "por caja", tiene precedencia y este valor se divide entre
# las cajas antes de usarse.
PAT_TOTAL = re.compile(
    r"(货箱重量总|总重量|total kg|总毛重|peso[_ ]?bruto|gross weight|\bg\.?w\.?\b)",
    re.I)
PAT_TOTAL_ALT = re.compile(r"(peso[_ ]?neto|净重|net weight)", re.I)
# Peso POR CAJA: exige la marca de "por caja" (单箱 / 单件 / por caja). Un
# patrón suelto como 毛重 o "gross weight" también casa con 总毛重, que es el
# TOTAL del renglón: eso daba pesos 10x (9.3 kg para un joyero de 0.93 kg).
PAT_CAJA = re.compile(
    r"(单箱毛重|单箱货箱\s*重量|单箱重量|单件重件|peso[_ ]?(?:por[_ ])?caja)", re.I)
# `ctns?` va SUELTO, sin el "total" delante, porque el encabezado no siempre
# dice "Total": PCIU9532241 trae `TotaI CTNS` con I MAYÚSCULA en vez de ele —
# un error de captura del proveedor que ninguna variante de "total ctn" empata.
# Buscar el token que sí es estable (CTN/CTNS) sobrevive a esos dedazos.
PAT_CAJAS = re.compile(
    r"(num_cajas|箱数|件数|\bctns?\b|cartons?|cajas)", re.I)
PAT_CBM = re.compile(r"(cbm_master|单箱体积|ctn measur|体积/箱|volumen/caja)", re.I)
PAT_PZCAJA = re.compile(r"(piezas_x_caja|单箱产品申报数量|产品申报数量|单箱数量|"
                        r"unit pcs|pcs/ctn|piezas por|per box|por caja)", re.I)
PAT_PZTOT = re.compile(
    r"(piezas_totales|产品总件数|总个数|总产品数量|total qty|total pcs)", re.I)
# Testigo del total REAL de la fila, que a veces contradice a la de arriba.
PAT_CANT = re.compile(r"(cantidad_total|总产品数量|total products)", re.I)
PAT_VALOR = re.compile(r"(valor_total|货值|total value|申报总价|total amount)", re.I)
PAT_VOLTOT = re.compile(r"(总体积|cbm_total|total volume|volumen total)", re.I)
PAT_L = re.compile(r"(长|largo|length)", re.I)
PAT_W = re.compile(r"(宽|ancho|width)", re.I)
PAT_H = re.compile(r"(高|alto|height)", re.I)
# `price` a secas al final: hay packing lists cuyo encabezado es literalmente
# "Price" (PCIU9532241). No se confunde con el importe de línea porque `col()`
# lo busca con `excluir_total=True`, que descarta cualquier "总"/"total".
PAT_PRECIO = re.compile(
    r"(单价|产品申报单价|precio_usd|unit price|u\.price|\bprice\b)", re.I)

_MAX_COLS_TEXTO = 14      # más allá de la 14 ya son notas y firmas
_MAX_TROZOS_TEXTO = 6


# ── dHash perceptual ─────────────────────────────────────────────────────────
def _abrir(datos: bytes):
    """Pillow → RGB, aplanando el alfa sobre BLANCO.

    Sin aplanar, una foto con fondo transparente se abre con el canal alfa
    puesto a negro y su hash no se parece en nada al de la misma foto guardada
    sobre blanco — que es como la guarda el packing list.
    """
    from PIL import Image

    im = Image.open(io.BytesIO(datos))
    if im.mode in ("RGBA", "LA", "P"):
        fondo = Image.new("RGB", im.size, (255, 255, 255))
        im = im.convert("RGBA")
        fondo.paste(im, mask=im.split()[-1])
        return fondo
    return im.convert("RGB")


# Lado máximo con el que se guarda cada foto del packing list. Sale del
# consumidor más exigente: la preparación para la IA (`_LADO_IA = 420` en
# packing_publicados). La UI las pide a 190 y 150, así que 420 le sobra a todos.
_LADO_GUARDADO = 420


def _reducir(datos: bytes, lado: int = _LADO_GUARDADO) -> bytes:
    """
    La foto encogida a ``lado`` px por el lado mayor, para no retenerla entera.

    Si Pillow no está o la imagen no se deja abrir, se devuelve el original: es
    preferible gastar memoria a quedarse sin la foto, que es el insumo del
    empate. Nunca AGRANDA — una foto ya chica se devuelve tal cual.
    """
    try:
        from PIL import Image
    except Exception:  # noqa: BLE001
        return datos
    try:
        with Image.open(io.BytesIO(datos)) as im:
            if max(im.size) <= lado:
                return datos
            im = im.convert("RGB")
            im.thumbnail((lado, lado), Image.LANCZOS)
            buf = io.BytesIO()
            im.save(buf, format="JPEG", quality=85)
            return buf.getvalue()
    except Exception:  # noqa: BLE001
        return datos


def dhash(datos: bytes, lado: int = 8) -> int | None:
    """
    Huella perceptual de 64 bits: gradiente horizontal sobre una miniatura 9×8.

    Es lo que rescata el empate cuando el sha256 no da porque los archivos no
    son idénticos byte a byte (Odoo recomprime, el proveedor recorta el margen).
    Compara cada píxel con su vecino DERECHO, así que sobrevive a cambios de
    brillo y de tamaño, no a un espejo ni a un recorte fuerte.

    Distancia entre dos huellas: ``bin(a ^ b).count("1")``. Devuelve ``None`` si
    la imagen no abre — un formato raro no debe tumbar el indexado del archivo.
    """
    try:
        from PIL import Image

        im = _abrir(datos).convert("L").resize(
            (lado + 1, lado), Image.Resampling.LANCZOS)
    except Exception:  # noqa: BLE001
        return None
    px = list(im.getdata())
    bits = 0
    for f in range(lado):
        base = f * (lado + 1)
        for x in range(lado):
            bits = (bits << 1) | (1 if px[base + x] > px[base + x + 1] else 0)
    return bits


def distancia(a: int, b: int) -> int:
    """Hamming entre dos dHash (0 = idénticas, 64 = opuestas)."""
    return bin(a ^ b).count("1")


# ── Índice ───────────────────────────────────────────────────────────────────
class Indice:
    """Un packing list ya leído: renglones, fotos hasheadas y números por caja.

    Se construye una vez por ARCHIVO y se comparte entre todos los SKUs que van
    a ese contenedor: parsear un xlsx de 100 MB con sus fotos cuesta segundos y
    varios SKUs suelen venir en el mismo embarque.
    """

    def __init__(self, datos: bytes, nombre: str, file_id: str = "") -> None:
        self.nombre = nombre
        self.file_id = file_id
        leido = packing_parser.leer(datos)
        self.filas: list[dict[str, Any]] = leido["filas"]
        self.avisos: list[str] = list(leido["avisos"])
        self.por_fila = {f.get("fila_excel"): f for f in self.filas}
        # fila_idx (0-based, la clave de las fotos) ↔ fila_excel (1-based, la
        # que ve el humano). Son distintos y confundirlos desplaza todo un renglón.
        self.idx_de_fila = {f.get("fila_idx"): f.get("fila_excel") for f in self.filas}
        self.fila_de_idx = {v: k for k, v in self.idx_de_fila.items()}

        self.fotos: dict[int, dict[str, Any]] = {}
        for i, crudo in (leido["imagenes"] or {}).items():
            dh = dhash(crudo)
            if dh is None:
                continue
            # El sha256 y el dHash se calculan sobre el ORIGINAL —el peldaño 0
            # es empate EXACTO de archivo y reducir la imagen lo destruiría—,
            # pero lo que se GUARDA es una copia reducida: los tres consumidores
            # (miniatura de la UI, foto del candidato y preparación para la IA)
            # la vuelven a achicar de todos modos, así que retener los bytes
            # originales solo cuesta RAM.
            #
            # Y costaba mucha: un índice vive hasta 3 h y caben 6 trabajos a la
            # vez. Medido, dos packing lists retenían 112 MB solo de fotos
            # (65.1 MB de uno de 160 renglones y 46.9 MB de otro de 167). Se
            # llegaba al GB con facilidad, en el mismo contenedor que atiende el
            # webhook de ventas de Mercado Libre.
            self.fotos[i] = {"sha": hashlib.sha256(crudo).hexdigest(),
                             "dh": dh, "crudo": _reducir(crudo)}

        wb = openpyxl.load_workbook(io.BytesIO(datos), data_only=True)
        # La MISMA hoja que leyó el parser. Si aquí se tomara sheetnames[0] se
        # estarían indexando dos hojas distintas del mismo libro.
        self.ws = wb.active
        self._cajas_grupo: dict[int, float | None] = {}
        self._mapear_columnas()
        self._precalcular_grupos()

    # ── Encabezado y columnas ────────────────────────────────────────────────
    def _mapear_columnas(self) -> None:
        todas = list(self.ws.iter_rows(values_only=True))
        h_idx = packing_parser.encontrar_encabezado(todas) if todas else 0
        self.fila_encabezado = h_idx + 1        # 1-based, como openpyxl
        enc = [str(v or "") for v in (todas[h_idx] if todas else [])]

        def col(patron: re.Pattern, excluir_total: bool = False) -> int | None:
            for i, e in enumerate(enc):
                if not patron.search(e):
                    continue
                if excluir_total and ("总" in e or "total" in e.lower()):
                    continue
                return i + 1                     # openpyxl cuenta desde 1
            return None

        self.c_total = col(PAT_TOTAL) or col(PAT_TOTAL_ALT)
        self.c_caja = col(PAT_CAJA)
        self.c_cajas = col(PAT_CAJAS)
        self.c_cbm = col(PAT_CBM)
        self.c_pzcaja = col(PAT_PZCAJA)
        self.c_pztot = col(PAT_PZTOT)
        # El precio UNITARIO nunca lleva 总/total: eso es el importe de la línea,
        # y confundirlos multiplica el costo del producto por sus piezas.
        self.c_precio = col(PAT_PRECIO, excluir_total=True)
        self.c_cant = col(PAT_CANT)
        self.c_valor = col(PAT_VALOR)
        self.c_voltot = col(PAT_VOLTOT)
        self.c_l, self.c_w, self.c_h = col(PAT_L), col(PAT_W), col(PAT_H)

        # Los avisos que vienen de `packing_parser` hablan de SUS columnas, y
        # este módulo tiene su propio juego de patrones —más amplio, porque lo
        # fue ganando archivo por archivo—. Cuando aquí sí se encontró lo que
        # allá faltó, el aviso del parser es FALSO y hay que callarlo: decirle
        # al usuario "este archivo no trae precio" mientras se le calcula el
        # costo CON el precio lo hace desconfiar del número correcto.
        # (PCIU9532241: el parser no reconoce `Price`, `TotaI CTNS` ni
        # `Quantity per box`; este módulo sí.)
        _resueltos: list[tuple[Any, str]] = [
            (self.c_precio, "precio unitario"),
            (self.c_pztot or self.c_pzcaja, "piezas totales"),
            (self.c_cbm or self.c_voltot or (self.c_l and self.c_w and self.c_h), "CBM"),
        ]
        for valor, marca in _resueltos:
            if valor:
                self.avisos = [a for a in self.avisos if marca.lower() not in a.lower()]

        if not self.c_precio:
            self.avisos.append(
                f"{self.nombre}: no se detectó columna de precio unitario USD "
                "(packing list puro): el costo de producto se conserva y solo "
                "se recalcula el flete.")
        if not (self.c_cbm or self.c_voltot or (self.c_l and self.c_w and self.c_h)):
            self.avisos.append(
                f"{self.nombre}: no se detectó ni volumen ni medidas de caja; "
                "el flete de sus renglones sale en CERO.")

    def _precalcular_grupos(self) -> None:
        """``{fila_excel: [filas del cartón]}`` en UNA pasada por los merges.

        Una celda de volumen fusionada sobre tres filas dice que esos tres
        productos viajan en la misma caja master, y entonces el flete se reparte
        entre las piezas de las tres — por pieza, en partes iguales.

        Se mira la columna de CBM y, si el archivo no la trae, la de cajas: sin
        ese respaldo un packing list sin columna de volumen simplemente no
        detectaba cajas mixtas y cada renglón se llevaba el cartón entero.
        """
        self.grupos: dict[int, list[int]] = {}
        col = self.c_cbm or self.c_cajas
        if not col:
            return
        for mr in self.ws.merged_cells.ranges:
            if mr.min_col <= col <= mr.max_col and mr.max_row > mr.min_row:
                filas = list(range(mr.min_row, mr.max_row + 1))
                for r in filas:
                    self.grupos[r] = filas

    def grupo(self, fila_excel: int) -> list[int]:
        return self.grupos.get(fila_excel, [fila_excel])

    def cajas(self, fila_excel: int) -> float | None:
        """
        Cuántas cajas trae el CARTÓN de esta fila, heredadas del ancla del grupo.

        El número de cajas es propiedad del cartón, no del renglón: cuando dos
        productos comparten cartón la celda está FUSIONADA y openpyxl solo la
        contesta en la fila ancla — en las de continuación devuelve ``None``.
        Leerla fila por fila y suplir ese ``None`` con un 1 le atribuye a la
        fila de continuación su TOTAL de renglón como si fueran las piezas de
        UNA caja: ``piezas_grupo`` se dispara y el flete por pieza se desploma
        (medido: 4,016 piezas donde eran 116 → $0.50 de flete donde eran
        $17.21). Es el gemelo silencioso del bug del $187.69: aquí SUB-costea,
        y un costo bajo se convierte en un precio bajo.

        ``None`` significa "el archivo no lo dice" y NO se traduce a 1: esa
        suposición es justamente la que produce el error mudo. Quien la use
        para dividir tiene que marcar el renglón como no confiable.
        """
        g = self.grupo(fila_excel)
        ancla = g[0]                      # el ancla es la fila de arriba del merge
        if ancla not in self._cajas_grupo:
            # Se resuelve UNA vez por cartón y se memoriza para todas sus filas:
            # `datos` lo pide una vez por fila del grupo.
            v = next((n for r in g if (n := self._num(r, self.c_cajas))), 0.0)
            for r in g:
                self._cajas_grupo[r] = v or None
        return self._cajas_grupo[ancla]

    # ── Lectura de celdas ────────────────────────────────────────────────────
    def _num(self, fila: int, col: int | None) -> float:
        if not col:
            return 0.0
        try:
            return float(self.ws.cell(row=fila, column=col).value)
        except (TypeError, ValueError):
            return 0.0

    def texto_fila(self, fila_excel: int) -> str:
        """
        Texto CRUDO de la fila — lo que se le da a leer al modelo.

        Crudo a propósito: hay archivos "normalizados" donde la columna rotulada
        `producto` trae en realidad el MATERIAL (MRKU3436938: 'plastics', 'EVA',
        'nylon'). El parser le cree al encabezado; para reconocer un producto es
        más seguro darle todo el renglón tal como viene.
        """
        vals: list[str] = []
        tope = min(self.ws.max_column or 0, _MAX_COLS_TEXTO)
        for c in range(1, tope + 1):
            v = self.ws.cell(row=fila_excel, column=c).value
            if not isinstance(v, str):
                continue
            t = v.strip()[:44]
            if len(t) > 1 and t not in vals and not t.replace(".", "").isdigit():
                vals.append(t)
        return " | ".join(vals[:_MAX_TROZOS_TEXTO])

    # ── Números POR CAJA ─────────────────────────────────────────────────────
    def _total_fila(self, fila: int) -> float:
        """
        Piezas TOTALES del renglón, con el importe como testigo principal.

        ``valor_total / precio_usd`` es el dato más confiable del archivo:
        son dos columnas independientes que solo cuadran si las dos dicen la
        verdad. Cuando falta, se recurre a las columnas de cantidad, y ahí entra
        la trampa clásica: si ``piezas_totales`` coincide con las piezas POR
        CAJA declaradas, esa columna eran piezas por caja y el total real es el
        producto por el número de cajas.
        """
        precio, valor = self._num(fila, self.c_precio), self._num(fila, self.c_valor)
        if precio > 0 and valor > 0:
            return valor / precio
        if cant := self._num(fila, self.c_cant):
            return cant
        pt = self._num(fila, self.c_pztot)
        cajas = self.cajas(fila) or 0.0
        pc = self._num(fila, self.c_pzcaja)
        if pt and cajas and pc and abs(pt - pc) < 0.01:
            return pt * cajas
        return pt or (pc * cajas if pc and cajas else 0.0)

    def _piezas_en_caja(self, fila: int) -> float:
        # El 1 de respaldo es solo para no dividir entre cero: cuando las cajas
        # son desconocidas el renglón sale marcado como no confiable en
        # :meth:`datos`, que es lo que impide que la UI lo apruebe solo.
        return self._total_fila(fila) / (self.cajas(fila) or 1)

    def datos(self, fila_excel: int) -> dict[str, Any]:
        """
        Los números de un renglón, emparejados POR CAJA.

        El orden de las operaciones ES la fórmula:

            total_fila     = valor_total/precio → cantidad_total → piezas×cajas
            piezas_en_caja = total_fila / num_cajas   (num_cajas del ANCLA)
            piezas_grupo   = Σ piezas_en_caja del cartón compartido
            cbm_por_pieza  = cbm_caja  / piezas_grupo
            peso_pieza     = peso_caja / piezas_grupo

        Dividir el CBM TOTAL del renglón entre las piezas de UNA caja es el bug
        que multiplica el flete por el número de cajas.
        """
        g = self.grupo(fila_excel)
        cajas_g = self.cajas(fila_excel)
        piezas_g = sum(self._piezas_en_caja(r) for r in g)

        # CBM de LA CAJA, en orden de confianza: columna propia → volumen total
        # del renglón entre sus cajas → las medidas.
        cbm_caja = next((v for r in g if (v := self._num(r, self.c_cbm))), 0.0)
        L = next((v for r in g if (v := self._num(r, self.c_l))), 0.0)
        W = next((v for r in g if (v := self._num(r, self.c_w))), 0.0)
        H = next((v for r in g if (v := self._num(r, self.c_h))), 0.0)
        origen_cbm = "columna_caja" if cbm_caja else ""
        if not cbm_caja and cajas_g:
            for r in g:
                if vt := self._num(r, self.c_voltot):
                    cbm_caja, origen_cbm = vt / cajas_g, "volumen_total"
                    break
        if not cbm_caja and L and W and H:
            cbm_caja, origen_cbm = L * W * H / 1_000_000, "medidas"

        # Peso de LA CAJA. Si solo hay columna del total del renglón, se divide
        # entre sus cajas (nunca al revés: multiplicar infla el peso 10x).
        peso_caja = next((v for r in g if (v := self._num(r, self.c_caja))), 0.0)
        if not peso_caja:
            for r in g:
                if v := self._num(r, self.c_total):
                    peso_caja = v / (cajas_g or 1)
                    break

        pieza_lwh = (packing_costos.dims_pieza(L, W, H, piezas_g)
                     if (L and W and H) else (0.0, 0.0, 0.0))
        return {
            "fila_excel": fila_excel,
            "grupo": [x for x in g if x != fila_excel],
            "piezas_grupo": piezas_g,
            "piezas_fila": self._total_fila(fila_excel),
            "cajas": cajas_g,
            # Sin cajas no hay "por caja": las piezas del grupo, el flete por
            # pieza y el peso por pieza salen de una división supuesta. Se
            # calculan igual para que el humano vea el renglón, pero el costo
            # no puede aprobarse solo con ellos.
            "confiable": cajas_g is not None,
            "motivo": ("" if cajas_g is not None else
                       "el packing list no dice cuántas cajas trae este cartón: "
                       "las piezas por caja y el flete son una suposición"),
            "caja_lwh": (L, W, H),
            "pieza_lwh": pieza_lwh,
            "cbm_caja": cbm_caja,
            "cbm_origen": origen_cbm or "sin_datos",
            "cbm_por_pieza": (cbm_caja / piezas_g) if piezas_g else 0.0,
            "peso_total": peso_caja,
            "peso_pieza": (peso_caja / piezas_g) if piezas_g else 0.0,
            "precio_usd": self._num(fila_excel, self.c_precio),
        }

    # ── Diagnóstico ──────────────────────────────────────────────────────────
    @property
    def n(self) -> int:
        return len(self.filas)

    @property
    def con_precio(self) -> int:
        """Cuántos renglones traen precio unitario. Cero = packing list PURO
        (sin factura), y entonces el costo de producto NO se recalcula."""
        if not self.c_precio:
            return 0
        return sum(1 for f in self.filas
                   if self._num(f.get("fila_excel") or 0, self.c_precio) > 0)


def indexar(datos: bytes, nombre: str, file_id: str = "") -> Indice:
    """Atajo legible; toda la lógica vive en :class:`Indice`."""
    ix = Indice(datos, nombre, file_id)
    log.info("packing list %s: %d renglones · %d fotos · precio en %d",
             nombre, ix.n, len(ix.fotos), ix.con_precio)
    return ix


def costo_de(dd: dict[str, Any], tarifa_mxn_m3: float, tipo_cambio: float,
             costo_producto_guardado: float | None = None,
             ) -> dict[str, Any]:
    """
    Costo por pieza a partir de los números por caja. Fórmula avalada por
    Brandon el 21-ago-2026:

        flete    = cbm_por_pieza × 7,500 MXN/m³      (tarifa FIJA)
        producto = precio_usd × 19
        costo    = producto + flete

    La tarifa es FIJA a propósito y difiere de ``packing_costos.calcular``, que
    prorratea 525,000 MXN entre el CBM del archivo: ese prorrateo necesita el
    contenedor COMPLETO y aquí hay uno o unos pocos SKUs. Con 70 m³ los dos dan
    lo mismo; con un archivo parcial, no.

    Packing list PURO (sin columna de precio): el costo de producto NO se
    inventa ni se pone en cero — se conserva el que ya está en kubera y se dice
    de dónde salió en ``origen_prod``. Solo el flete se rehace.

    Y si NINGUNA de las dos fuentes lo trae, la fila sale **incompleta**
    (``completo=False``, ``costo=None``) con su motivo. No hay tercera opción:
    tratar el "no hay nada" como un cero escribía un costo que era SOLO FLETE
    —etiquetado ``origen_prod='kubera'``, o sea afirmando que se respetó un
    costo previo que nunca existió— con confianza suficiente para que la UI lo
    aprobara solo y le pusiera encima el candado de COSTO VALIDADO. Un costo
    que es puro flete es un precio de venta por debajo de lo que costó la
    mercancía.
    """
    cbm_pz = float(dd.get("cbm_por_pieza") or 0.0)
    precio = float(dd.get("precio_usd") or 0.0)
    flete = cbm_pz * tarifa_mxn_m3
    faltas: list[str] = []
    if not dd.get("confiable", True):
        faltas.append(dd.get("motivo") or "el renglón no es confiable")

    # `None` es "no había costo" y 0 es "el costo guardado es cero": ninguno de
    # los dos es un costo de producto, pero se distinguen para que el motivo
    # diga la verdad y el humano sepa qué capturar.
    guardado = (None if costo_producto_guardado is None
                else float(costo_producto_guardado))
    if precio > 0:
        producto, origen = precio * tipo_cambio, "packing_list"
    elif guardado and guardado > 0:
        producto, origen = guardado, "kubera"
    else:
        producto, origen = None, None
        faltas.append(
            "el packing list no trae precio y "
            + ("no hay costo guardado en kubera" if guardado is None
               else "el costo guardado es 0")
            + ": captúralo a mano")

    completo = not faltas
    return {"flete": round(flete, 4),
            "producto_mxn": None if producto is None else round(producto, 4),
            "origen_prod": origen,
            "costo": (round(producto + flete, 4)
                      if completo and producto is not None else None),
            "completo": completo, "motivo": " · ".join(faltas)}
