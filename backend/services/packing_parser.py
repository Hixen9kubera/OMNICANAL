"""
packing_parser.py — Lee un packing list chino (.xlsx) y devuelve filas + fotos.

Portado de kubera/costos/app.py (validado contra ~60 contenedores reales). Dos
cosas que NO hay que reinventar:

1. **openpyxl no ve las imágenes embebidas** (``ws._images`` sale vacío cuando el
   ancla es ``oneCellAnchor``). Hay que abrir el .xlsx como ZIP y cruzar a mano
   ``xl/drawings/drawingN.xml`` (fila/columna del ancla) con
   ``xl/drawings/_rels/drawingN.xml.rels`` (rId → ``xl/media/imageN.ext``).
2. **Las cajas compartidas se detectan por celdas MERGED** en la columna de
   volumen: si el merge abarca 3 filas, esos 3 SKUs viajan en la misma caja
   master y hay que repartir su CBM.

CBM por pieza — en orden de prioridad:
  a) Columna "Total Volume" (总体积) de la fila ÷ piezas de la fila. Es el dato
     autoritativo cuando el packing list lo trae.
  b) Caja compartida: cbm_caja ÷ piezas_totales_de_la_caja.
  c) Caja propia: cbm_master ÷ piezas_por_caja  (caso (b) con un solo SKU).
  d) Sin volumen declarado: L×W×H/1e6 ÷ piezas_por_caja.
"""
from __future__ import annotations

import io
import logging
import re
import zipfile
from collections import Counter
from typing import Any
from xml.etree import ElementTree as ET

import openpyxl

log = logging.getLogger("omnicanal.packing.parser")

NS = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "rels": "http://schemas.openxmlformats.org/package/2006/relationships",
}


# ── Imágenes embebidas ───────────────────────────────────────────────────────
def _rutas_media(target: str) -> list[str]:
    """
    Candidatos de ruta REAL dentro del ZIP para el ``Target`` de un ``.rels``.

    Los generadores no se ponen de acuerdo y la especificación permite las tres
    formas. Excel escribe ``../media/image1.png`` (relativo a ``xl/drawings/``);
    **openpyxl y WPS Office —el que usa media China— escriben
    ``/xl/media/image1.png``, absoluto con diagonal inicial**; y algunos dejan
    ``media/image1.png`` a secas. Los nombres dentro del ZIP nunca llevan la
    diagonal, así que la forma absoluta no empata y el ancla se descartaba EN
    SILENCIO: el packing list "no traía fotos" aunque las trajera, y con eso
    muerto la etapa de empate por imagen nunca corría.

    Se devuelven candidatos en vez de una ruta porque resolver ``media/x`` es
    ambiguo (relativo a la carpeta del drawing o al paquete): quien llama se
    queda con el primero que exista de verdad en el ZIP.
    """
    t = target.replace("\\", "/").strip().lstrip("/")
    if t.startswith("../"):
        t = t[3:]
        return [f"xl/{t}"]
    if t.startswith("xl/"):
        return [t]
    # Relativo a xl/drawings/ según la especificación, pero en la práctica casi
    # siempre quiere decir xl/media/. Se prueban las dos.
    return [f"xl/drawings/{t}", f"xl/{t}"]


def extraer_imagenes(xlsx_bytes: bytes, columna: int | None = None) -> dict[int, bytes]:
    """
    Devuelve ``{fila_0based: bytes}`` con la foto de producto de cada fila.

    Los packing lists ponen la foto en columnas distintas según el proveedor, así
    que por defecto se autodetecta: la columna con MÁS anclas es la de producto
    (las otras suelen ser el logo del proveedor o sellos, 1-2 imágenes sueltas).
    """
    out: dict[int, bytes] = {}
    sin_resolver = 0
    with zipfile.ZipFile(io.BytesIO(xlsx_bytes)) as z:
        nombres = set(z.namelist())
        # Puede haber varios drawings (uno por hoja); recorremos todos.
        drawings = sorted(n for n in nombres
                          if re.fullmatch(r"xl/drawings/drawing\d+\.xml", n))
        anclas: list[tuple[int, int, str]] = []  # (col, fila, ruta_media)
        for draw in drawings:
            rels_path = f"xl/drawings/_rels/{draw.split('/')[-1]}.rels"
            if rels_path not in nombres:
                continue
            try:
                rels = {}
                for r in (ET.fromstring(z.read(rels_path).decode("utf-8"))
                            .findall("rels:Relationship", NS)):
                    rels[r.attrib["Id"]] = next(
                        (c for c in _rutas_media(r.attrib["Target"]) if c in nombres),
                        None)
                root = ET.fromstring(z.read(draw).decode("utf-8"))
            except Exception as exc:  # noqa: BLE001
                log.warning("drawing %s ilegible: %s", draw, exc)
                continue

            for a in (root.findall("xdr:twoCellAnchor", NS)
                      + root.findall("xdr:oneCellAnchor", NS)):
                frm = a.find("xdr:from", NS)
                blip = a.find(".//a:blip", NS)
                if frm is None or blip is None:
                    continue
                col_el, row_el = frm.find("xdr:col", NS), frm.find("xdr:row", NS)
                if col_el is None or row_el is None:
                    continue
                rid = blip.attrib.get(f"{{{NS['r']}}}embed")
                destino = rels.get(rid)
                if destino:
                    anclas.append((int(col_el.text), int(row_el.text), destino))
                else:
                    # Nunca en silencio: un ancla sin media resoluble es la
                    # diferencia entre "el proveedor no mandó fotos" y "no
                    # supimos leerlas".
                    sin_resolver += 1

        if columna is None and anclas:
            columna = Counter(col for col, _, _ in anclas).most_common(1)[0][0]

        for col, fila, media in anclas:
            if columna is not None and col != columna:
                continue
            try:
                out[fila] = z.read(media)
            except Exception as exc:  # noqa: BLE001
                log.warning("media %s ilegible: %s", media, exc)

    if sin_resolver:
        log.warning("packing list: %s anclas de imagen sin media resoluble "
                    "(Target del .rels en forma inesperada)", sin_resolver)
    log.info("imágenes embebidas: %s anclas → %s filas con foto (columna %s)",
             len(anclas), len(out), columna)
    return out


def tiene_media(xlsx_bytes: bytes) -> bool:
    """¿El .xlsx trae imágenes embebidas, sin importar si supimos anclarlas?

    Sirve para distinguir "el proveedor no mandó fotos" de "no las supimos
    leer" en los avisos de :func:`leer`.
    """
    try:
        with zipfile.ZipFile(io.BytesIO(xlsx_bytes)) as z:
            return any(n.startswith("xl/media/") for n in z.namelist())
    except Exception:  # noqa: BLE001
        return False


# ── Encabezados ──────────────────────────────────────────────────────────────
def _norm(h: Any) -> str:
    return str(h).strip().lower().replace("\n", " ") if h else ""


def mapear_columnas(header: list[Any]) -> dict[str, int]:
    """
    Mapea el encabezado (chino / inglés / español) a nombres internos. Los
    packing lists varían muchísimo entre proveedores; estos patrones salieron de
    procesar los contenedores reales del Drive.
    """
    cm: dict[str, int] = {}
    for i, h in enumerate(header):
        hl = _norm(h)
        if not hl:
            continue

        # Descripción del producto (inglés preferido, chino como respaldo)
        if ("英文" in hl or "english" in hl or "description of goods" in hl
                or hl in ("des.", "description", "descripcion", "descripción")):
            cm.setdefault("producto", i)
        elif "producto" not in cm and ("中文" in hl or "chino" in hl):
            cm["producto"] = i
        if "中文" in hl or "chino" in hl:
            cm.setdefault("producto_chn", i)

        # Cantidades
        if any(k in hl for k in ("箱数", "ctns", "total ctn", "cantidad fisica",
                                 "num_cajas", "caja", "ctn")):
            cm.setdefault("cajas", i)
        if any(k in hl for k in ("总个数", "total qty", "qty(pcs)", "piezas_totales",
                                 "总数量", "t.qty", "总产品数量")):
            cm.setdefault("piezas_total", i)
        elif hl == "qty" and "piezas_total" not in cm:
            cm["piezas_total"] = i

        # Dimensiones de la caja master
        if any(k in hl for k in ("长(cm)", "length", "largo", "货箱长度")):
            cm.setdefault("largo", i)
        if any(k in hl for k in ("宽(cm)", "width", "ancho", "货箱宽度")):
            cm.setdefault("ancho", i)
        if any(k in hl for k in ("高(cm)", "height", "altura", "货箱高度", "alto")):
            cm.setdefault("alto", i)

        # Volumen TOTAL de la fila — autoritativo cuando existe.
        if any(k in hl for k in ("总体积", "total volume", "total cbm", "cbm total",
                                 "volumen total", "总立方")):
            cm.setdefault("cbm_total_fila", i)
        # Volumen por caja master
        elif any(k in hl for k in ("体积/箱", "@cbm", "ctn measurement", "单箱体积",
                                   "cbm_master_cart", "cbm_master_carton",
                                   "volumen/caja", "meas")):
            cm.setdefault("cbm_master", i)
        elif hl == "cbm" and "cbm_master" not in cm:
            cm["cbm_master"] = i

        # Precio unitario (excluye los "total" / 货值 que son importes de línea)
        if any(k in hl for k in ("单价", "unit price", "precio_usd", "precio/usd",
                                 "u.price", "precio unitario")):
            if "total" not in hl and "货值" not in hl:
                cm.setdefault("precio_usd", i)

        # Cantidad TOTAL de producto del renglón. Ojo: muchos proveedores llaman
        # "piezas_totales" a las piezas POR CAJA y ponen el total real en otra
        # columna. Se capturan ambas y el detector decide cuál es cuál.
        if any(k in hl for k in ("cantidad_total_productos", "cantidad total",
                                 "总产品数量", "total products", "total pcs")):
            cm.setdefault("cantidad_total", i)
        # Importe de la línea: valor_total / precio_usd delata el total de piezas.
        if any(k in hl for k in ("valor_total", "货值", "total value", "amount",
                                 "total amount")):
            cm.setdefault("valor_total", i)

        # Peso
        if any(k in hl for k in ("总毛重", "gross weight", "毛重", "g.w")):
            cm.setdefault("peso_bruto", i)
        elif any(k in hl for k in ("总净重", "net weight", "净重", "n.w")) \
                and "peso_bruto" not in cm:
            cm["peso_bruto"] = i
    return cm


def encontrar_encabezado(filas: list[tuple], max_fila: int = 15) -> int:
    """
    Índice (0-based) de la fila de encabezado: la primera con al menos 5 celdas
    llenas y algún texto. Los packing lists traen 2-6 filas de membrete arriba.
    """
    mejor_idx, mejor_score = 0, -1
    for i, fila in enumerate(filas[:max_fila]):
        llenas = [v for v in fila if v not in (None, "")]
        if len(llenas) < 5:
            continue
        textos = [v for v in llenas if isinstance(v, str) and len(v) > 1]
        # Gana la fila que mapea MÁS columnas conocidas, no la primera que pasa
        # un umbral: hay membretes que traen 5-6 celdas y engañan al corte rápido.
        score = len(mapear_columnas(list(fila)))
        if textos and score > mejor_score:
            mejor_idx, mejor_score = i, score
    return mejor_idx


def _f(v: Any) -> float:
    """Float tolerante: '12.5 cm', '1,234', None → 12.5 / 1234.0 / 0.0."""
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    txt = str(v).replace(",", "").strip()
    m = re.search(r"-?\d+(?:\.\d+)?", txt)
    return float(m.group(0)) if m else 0.0


# ── Cajas compartidas (celdas merged) ────────────────────────────────────────
def _grupos_por_merge(ws, col_1based: int) -> dict[int, tuple[int, int]]:
    """
    ``{fila_excel: (fila_cabecera, tamaño_grupo)}`` para los merges que cruzan la
    columna dada. Un merge de 3 filas = 3 SKUs compartiendo una caja master.
    """
    grupos: dict[int, tuple[int, int]] = {}
    for mr in ws.merged_cells.ranges:
        if mr.min_col <= col_1based <= mr.max_col and mr.max_row > mr.min_row:
            tam = mr.max_row - mr.min_row + 1
            for r in range(mr.min_row, mr.max_row + 1):
                grupos[r] = (mr.min_row, tam)
    return grupos


# ── Lectura completa ─────────────────────────────────────────────────────────
def leer(xlsx_bytes: bytes, columna_imagen: int | None = None) -> dict[str, Any]:
    """
    Devuelve::

        {
          "columnas":   {nombre_interno: indice},     # lo que se detectó
          "encabezado": 3,                            # fila 0-based
          "filas":      [ {...}, ... ],               # una por fila con producto
          "imagenes":   {fila_0based: bytes},
          "avisos":     ["..."],
        }

    Cada fila trae los datos crudos y ``cbm_por_pieza`` ya resuelto, con
    ``cbm_origen`` indicando de dónde salió (para poder auditarlo en la UI).
    """
    avisos: list[str] = []
    imagenes = extraer_imagenes(xlsx_bytes, columna_imagen)

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active
    todas = list(ws.iter_rows(values_only=True))
    if not todas:
        raise ValueError("El archivo no tiene filas.")

    h_idx = encontrar_encabezado(todas)
    cm = mapear_columnas(list(todas[h_idx]))
    if "producto" not in cm:
        raise ValueError(
            "No se encontró la columna de descripción del producto "
            f"(encabezado detectado en la fila {h_idx + 1})."
        )
    if "piezas_total" not in cm:
        avisos.append("No se detectó columna de piezas totales; las cantidades "
                      "quedan en 0 y hay que capturarlas a mano.")

    # Los merges se miran en la columna de volumen; si no hay, en la de cajas.
    col_merge = cm.get("cbm_master", cm.get("cajas"))
    grupos = _grupos_por_merge(ws, col_merge + 1) if col_merge is not None else {}

    filas: list[dict[str, Any]] = []
    for idx in range(h_idx + 1, len(todas)):
        fila = todas[idx]
        if not any(v not in (None, "") for v in fila):
            continue
        if cm["producto"] >= len(fila) or not fila[cm["producto"]]:
            continue

        fila_excel = idx + 1
        cab_excel, tam_grupo = grupos.get(fila_excel, (fila_excel, 1))
        cab = todas[cab_excel - 1]

        def _cab(clave: str) -> float:
            """Los atributos de la caja master solo están en la fila cabecera."""
            i = cm.get(clave)
            return _f(cab[i]) if i is not None and i < len(cab) else 0.0

        def _fil(clave: str) -> float:
            i = cm.get(clave)
            return _f(fila[i]) if i is not None and i < len(fila) else 0.0

        cajas = _cab("cajas")
        cbm_master = _cab("cbm_master")
        largo, ancho, alto = _cab("largo"), _cab("ancho"), _cab("alto")
        peso_caja = (_cab("peso_bruto") / cajas) if cajas > 0 else 0.0
        piezas = _fil("piezas_total")

        # Piezas de TODAS las filas del grupo (las que comparten la caja).
        piezas_grupo = sum(
            _f(todas[cab_excel - 1 + k][cm["piezas_total"]])
            for k in range(tam_grupo)
            if "piezas_total" in cm
            and cab_excel - 1 + k < len(todas)
            and cm["piezas_total"] < len(todas[cab_excel - 1 + k])
        ) or piezas

        # Volumen del grupo completo (todas las cajas master de esa línea).
        if cbm_master <= 0 and largo and ancho and alto:
            cbm_master = (largo * ancho * alto) / 1_000_000
            origen_master = "dimensiones"
        else:
            origen_master = "declarado"
        cbm_grupo = cbm_master * cajas if cajas > 0 else cbm_master

        # ── CBM por pieza, en orden de prioridad ──
        cbm_total_fila = _fil("cbm_total_fila")
        if cbm_total_fila > 0 and piezas > 0:
            cbm_por_pieza = cbm_total_fila / piezas
            cbm_origen = "total_volume"
        elif cbm_grupo > 0 and piezas_grupo > 0:
            # Reparte el volumen del grupo entre TODAS sus piezas: cada pieza
            # ocupa lo mismo, compartan caja o no.
            cbm_por_pieza = cbm_grupo / piezas_grupo
            cbm_origen = "caja_compartida" if tam_grupo > 1 else "caja_propia"
        else:
            cbm_por_pieza = 0.0
            cbm_origen = "sin_datos"

        producto = str(fila[cm["producto"]]).strip()
        chn = ""
        if "producto_chn" in cm and cm["producto_chn"] < len(fila) and fila[cm["producto_chn"]]:
            chn = str(fila[cm["producto_chn"]]).strip()

        filas.append({
            "fila_excel": fila_excel,
            "fila_idx": idx,                     # 0-based: la clave de `imagenes`
            "producto": producto,
            "producto_chn": chn,
            "cajas": cajas,
            "piezas": piezas,
            "piezas_grupo": piezas_grupo,
            "comparte_caja": tam_grupo > 1,
            "tam_grupo": tam_grupo,
            "largo": largo,
            "ancho": ancho,
            "alto": alto,
            "peso_caja": peso_caja,
            "peso_pieza": (peso_caja / (piezas_grupo / cajas)) if cajas > 0 and piezas_grupo > 0 else 0.0,
            "cbm_master": cbm_master,
            "cbm_master_origen": origen_master,
            "cbm_total_fila": cbm_total_fila,
            "cbm_por_pieza": cbm_por_pieza,
            "cbm_origen": cbm_origen,
            "precio_usd": _fil("precio_usd"),
            # Para el detector de semántica (ver packing_costos.normalizar_semantica)
            "cantidad_total": _fil("cantidad_total"),
            "valor_total": _fil("valor_total"),
            "peso_bruto_fila": _cab("peso_bruto"),
            "imagen": idx in imagenes,
        })

    if not filas:
        raise ValueError("No se encontró ninguna fila de producto bajo el encabezado.")

    sin_cbm = sum(1 for f in filas if f["cbm_por_pieza"] <= 0)
    if sin_cbm:
        avisos.append(f"{sin_cbm} de {len(filas)} filas quedaron sin CBM por pieza.")
    sin_img = sum(1 for f in filas if not f["imagen"])
    if sin_img == len(filas) and tiene_media(xlsx_bytes):
        # No es lo mismo que el proveedor no mande fotos: aquí SÍ vienen y no se
        # pudieron anclar a ningún renglón. Decirlo distinto, porque el empate
        # por imagen se queda sin insumo y la causa está en el archivo, no en él.
        avisos.append(
            f"El Excel SÍ trae imágenes embebidas, pero ninguna se pudo asociar "
            f"a un renglón (las {len(filas)} filas quedaron sin foto). El empate "
            f"por imagen no va a correr; revisar el ancla del drawing.")
    elif sin_img:
        avisos.append(f"{sin_img} de {len(filas)} filas no tienen foto en el Excel.")
    if not cm.get("precio_usd"):
        avisos.append("No se detectó columna de precio unitario USD.")

    return {
        "columnas": cm,
        "encabezado": h_idx,
        "filas": filas,
        "imagenes": imagenes,
        "avisos": avisos,
    }


def contenedor_desde_nombre(nombre_archivo: str) -> str:
    """
    Extrae el código de contenedor del nombre del archivo:
    ``MRSU6548934``, ``PRY25-650``, ``149504112342``. Portado de costos/app.py.
    """
    base = re.sub(r"\.xlsx$", "", nombre_archivo, flags=re.IGNORECASE)
    limpio = re.sub(r"[一-鿿（）()]", "", base)
    limpio = re.sub(r"\(已瘦身\)", "", limpio, flags=re.IGNORECASE)
    if m := re.search(r"\b([A-Z]{4}\d{7,})\b", limpio.upper()):
        return m.group(1)
    if m := re.search(r"((?:PRY|LEX|TONY)\d{2}-\d{3,4})", limpio, re.IGNORECASE):
        return m.group(1).upper()
    if m := re.search(r"\b(\d{9,})\b", base):
        return m.group(1)
    tokens = re.split(r"[=\s_\-]", limpio.strip())
    return (tokens[0] if tokens and tokens[0] else base)[:20]
