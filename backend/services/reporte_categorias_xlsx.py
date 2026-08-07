"""
reporte_categorias_xlsx.py — El reporte ÚNICO de ventas y costos, generado EN
VIVO desde la BD kubera con los filtros elegidos en /analisis/reportes.

SIN MÁRGENES (Eduardo, 7-ago). El libro traía Ganancia y Margen %% en las tres
hojas y se retiraron: se calculaban sobre `costing.costos_validados` /
`costos_finales`, y esa base tiene defectos medidos que hacen el resultado no
publicable — ~30%% del catálogo con `costo_producto` puesto como precio USD×19
(placeholder, no costo medido), ~536 SKUs con el peso de la CAJA capturado como
peso de pieza, y 30 SKUs con `piezas_por_caja < 1` que multiplican el flete en
vez de dividirlo. Un margen sacado de ahí se ve como un hecho y no lo es. Las
columnas de costo SÍ se quedan: son el dato crudo, y quien lo necesite arma la
resta en su propia tabla dinámica sabiendo lo que está restando.

TRES hojas, que responden tres preguntas distintas (Eduardo, 5-ago):
  - Resumen:    ¿qué categoría vende? Una fila por categoría PRINCIPAL con SKUs
                con venta, unidades, venta $, costo base y costo final.
                Totales y %% con FÓRMULAS (recalcula solo).
  - Categorias: ¿qué SKU dentro de esa categoría? El árbol completo con
                subtotales SUBTOTAL(9,…) por nivel y, bajo cada hoja, sus
                publicaciones con las mismas columnas de costo.
  - Ventas:     ¿por qué? Una fila por línea vendida — el insumo para tablas
                dinámicas y para auditar una cifra que no cuadre.

FUENTE ÚNICA: los PEDIDOS (channel.orders/order_items). La comisión solo puede
salir de ahí, porque es donde vive la que REALMENTE cobró Mercado Libre;
mezclar fuentes haría que una columna dijera una venta y la de al lado restara
sobre otra. Consecuencia declarada: este libro NO cuadra con la página
/analisis/categorias, que sigue leyendo sales_daily_completa (esa serie incluye
el histórico rescatado de dailytrack). Sí cuadra con la pestaña VENTAS, que
también es 100%% pedidos.

Costo Base  = producto + flete de importación, × unidades vendidas
Costo Final = Costo Base + comisión REAL + envío estimado

Limitaciones que siguen vigentes: "Días en venta" no existe (listings no
guarda la fecha de alta → va la 1ª venta del período); los cargos de bodega
FULL quedan fuera (se facturan por mes, no por venta); Amazon reporta comisión
0 hasta tener Finances API; una fila sin costo capturado va en blanco, no en
cero.
"""
from __future__ import annotations

import io
from collections import defaultdict
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TIENDA = {"BEKURA": "Bekura", "SANCORFASHION": "Sancor", "AMAZON": "Amazon"}
_CAB_FILL = PatternFill("solid", fgColor="1F3864")
# Ámbar para la celda que falta o no es de fiar. Acompaña al texto de la
# columna Diagnóstico; nunca lo sustituye (el color no sobrevive a un CSV).
_AVISO_FILL = PatternFill("solid", fgColor="FFF2CC")
_NIVEL_FILL = ["BDD7EE", "DDEBF7", "F2F7FC", "FAFCFE"]
_MONEY, _INT, _PCT = "$#,##0", "#,##0", "0.0%"


def _f(**kw) -> Font:
    return Font(name="Arial", **{"size": 10, **kw})


class _Nodo:
    __slots__ = ("nombre", "hijos", "hojas")

    def __init__(self, nombre: str):
        self.nombre = nombre
        self.hijos: dict[str, _Nodo] = {}
        self.hojas: list[dict] = []  # hojas del árbol ML que caen aquí


def _costo_final(f: dict) -> float | None:
    """Costo Base + los cobros del canal. None cuando el costo del producto no
    está capturado: sin él la cifra sería el costo de vender algo gratis, y una
    celda vacía es más honesta que un costo final que finge estar completo."""
    if f.get("costo_base") is None:
        return None
    return (float(f["costo_base"]) + float(f.get("comision") or 0)
            + float(f.get("envio") or 0))


# ── DIAGNÓSTICO ──────────────────────────────────────────────────────────────
#
# Una celda vacía dice "no sé" sin decir QUÉ no sé, y un $0 miente peor: no
# distingue "el envío costó cero" de "no tenemos el dato". Esta columna nombra
# el problema de cada renglón y dice dónde se arregla.
#
# Va como TEXTO y no solo como color, porque el color no sobrevive a un copiar
# y pegar ni a una exportación a CSV, y este archivo se comparte fuera del
# panel. El relleno ámbar es la comodidad; el texto es el dato.
#
# Umbrales calibrados el 7-ago contra los 689 SKUs con venta en 60 días:
#   sin fila de costo .......  96 (13.9%)   ← los renglones vacíos
#   sin costo_fee_envio ..... 239 (34.7%)   ← lo que se pintaba como $0
#   costo placeholder ....... 202 (29.3%)
#   cajas = 0 ............... 151 (21.9%)
#   peso de caja ............  32 ( 4.6%)
#   piezas_por_caja < 1 .....   7 ( 1.0%)

# El peso de `costos_validados` está en KILOGRAMOS: verificado el 7-ago contra
# la báscula de ML (`ml_ficha.peso_g`, solo PACKAGE_WEIGHT) sobre 344
# publicaciones comparables — la mediana de peso×1000/peso_ML es 1.000.
# Arriba de esta densidad el "peso de la pieza" es en realidad el de la caja.
_DENSIDAD_MAX = 1.5          # kg/L

# Al capturar, ~30%% del catálogo trae el PRECIO en dólares multiplicado por un
# tipo de cambio redondo en vez de un costo medido. `fx_rate_used` está NULL en
# las 15,395 filas, así que el múltiplo exacto es la única huella que queda.
_TC_PLACEHOLDER = 19


def _num(v: Any) -> float | None:
    try:
        return None if v is None else float(v)
    except (TypeError, ValueError):
        return None


def diagnosticar(f: dict, ingreso: float | None = None,
                 costo_base: float | None = None) -> str:
    """Por qué esta fila está incompleta o no es de fiar. "" si está sana.

    Devuelve el problema MÁS GRAVE en claro, y si hay más los lista al final
    entre paréntesis: así se puede filtrar por el prefijo sin perder el resto.
    """
    graves: list[str] = []
    otros: list[str] = []

    tiene_fila = bool(f.get("tiene_validado") or f.get("tiene_final"))
    if costo_base is None:
        graves.append(
            "SIN COSTO — el SKU no está capturado en costing"
            + ("" if tiene_fila else " (no existe ni en costos_validados ni en "
                                    "costos_finales)")
            + "; sin eso no hay costo base ni costo final")

    ing, cb = _num(ingreso), _num(costo_base)
    if ing and cb and cb > ing:
        graves.append(f"COSTO MAYOR QUE LA VENTA — costo base ${cb:,.0f} contra "
                      f"un ingreso de ${ing:,.0f}")

    pzas = _num(f.get("piezas_por_caja"))
    if pzas is not None and 0 < pzas < 1:
        graves.append(f"FLETE MULTIPLICADO ×{1 / pzas:,.0f} — piezas_por_caja = "
                      f"{pzas:g}: al ser menor que 1, el flete de la caja se "
                      f"multiplica en vez de repartirse entre las piezas")

    peso, largo = _num(f.get("peso")), _num(f.get("largo"))
    alto, ancho = _num(f.get("alto")), _num(f.get("ancho"))
    if peso and largo and alto and ancho and largo > 0 and alto > 0 and ancho > 0:
        litros = largo * alto * ancho / 1000.0     # cm³ → L
        if litros > 0 and peso / litros > _DENSIDAD_MAX:
            graves.append(f"PESO DE CAJA — {peso / litros:,.1f} kg/L: el peso "
                          f"capturado ({peso:g} kg) parece el de la caja master, "
                          f"no el de una pieza; infla el costo de envío")

    cp = _num(f.get("costo_producto"))
    if cp and cp > 0 and round(cp * 100) % (_TC_PLACEHOLDER * 100) == 0:
        otros.append(f"COSTO PLACEHOLDER — costo_producto ${cp:,.0f} es múltiplo "
                     f"exacto de {_TC_PLACEHOLDER} (precio en USD por un tipo de "
                     f"cambio redondo), no un costo medido")

    if _num(f.get("cajas")) == 0:
        otros.append("CAJAS EN CERO — con cajas = 0 el costo por pieza no se "
                     "puede derivar del contenedor")

    if f.get("envio_origen") == "sin dato":
        otros.append("SIN DATO DE ENVÍO — ni ML nos dio el cobro del embarque "
                     "ni costos_finales tiene costo_fee_envio; la celda va "
                     "vacía, no en $0, y el costo final de este renglón queda "
                     "INCOMPLETO (es un piso: le falta el envío)")
    elif f.get("envio_origen") == "estimado":
        otros.append("ENVÍO ESTIMADO — es el cálculo por peso/dimensiones de "
                     "costing, no el cobro real de ML; el estimado ya se ha "
                     "equivocado por 4× en ambas direcciones")

    todos = graves + otros
    if not todos:
        return ""
    if len(todos) == 1:
        return todos[0]
    extra = " · ".join(t.split(" — ")[0] for t in todos[1:])
    return f"{todos[0]}  (además: {extra})"


def _arbol(hojas: list[dict]) -> tuple[dict[str, _Nodo], dict[str, dict]]:
    """Raíces del árbol + acumulados por nodo (uds/venta/costos/pubs/skus).

    `venta_con_costo` acumula aparte la venta de las filas que SÍ traen costo.
    Es el medidor de COBERTURA de la categoría: cuánto de lo que vendió tiene
    costo capturado. Si queda muy por debajo de `venta`, las columnas de costo
    de esa fila describen una muestra, no la categoría."""
    raices: dict[str, _Nodo] = {}
    acum: dict[str, dict] = defaultdict(lambda: {"uds": 0, "venta": 0.0,
                                                 "pubs": 0, "activas": 0,
                                                 "skus": 0, "costo_base": 0.0,
                                                 "costo_final": 0.0,
                                                 "venta_con_costo": 0.0})
    for h in hojas:
        partes = [p.strip() for p in str(h["ruta"]).split("›") if p.strip()] \
            or ["Sin categoría"]
        n = raices.setdefault(partes[0], _Nodo(partes[0]))
        clave = partes[0]
        for parte in partes[1:]:
            n = n.hijos.setdefault(parte, _Nodo(parte))
        n.hojas.append(h)
        for i in range(len(partes)):
            clave = "›".join(partes[: i + 1])
            a = acum[clave]
            a["uds"] += int(h["uds"] or 0)
            a["venta"] += float(h["venta"] or 0)
            a["pubs"] += int(h["publicaciones"] or 0)
            a["activas"] += int(h["activas"] or 0)
            a["skus"] += int(h["skus"] or 0)
            cf = _costo_final(h)
            if cf is not None:
                a["costo_base"] += float(h["costo_base"])
                a["costo_final"] += cf
                # la venta MEDIBLE viene calculada del SQL (solo los SKUs con
                # costo), no la venta total de la categoría
                a["venta_con_costo"] += float(h.get("venta_con_costo") or 0)
    return raices, acum


def construir(hojas: list[dict], pubs: list[dict], ventas: list[dict],
              desde: str, hasta: str, cuenta: str | None,
              censo_envio: dict[str, int] | None = None) -> bytes:
    """Arma el workbook y lo devuelve como bytes (para StreamingResponse)."""
    pubs_por_cat: dict[str, list[dict]] = defaultdict(list)
    for p in pubs:
        pubs_por_cat[str(p["category_id"])].append(p)
    raices, acum = _arbol(hojas)
    # "Sin categoría" SIEMPRE al final, venda lo que venda (Eduardo, 05-ago)
    orden_raiz = sorted(raices,
                        key=lambda r: (r == "Sin categoría", -acum[r]["venta"]))

    wb = Workbook()

    # ── hoja Categorias ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Categorias")
    # "Origen envío" y "Diagnóstico" van AL FINAL a propósito: así el bloque
    # numérico F..K queda contiguo y los SUBTOTAL(9,…) por nivel no tienen que
    # saltarse una columna de texto.
    cabs = ["Categoría / SKU", "Tienda", "Título", "MLM ID", "Situación",
            "Ventas Uds", "Ventas $", "Costo base", "Comisión ML", "Envío",
            "Costo final", "Precio", "1ª venta", "Últ. venta", "Origen envío",
            "Diagnóstico"]
    _N = len(cabs)
    for c, t in enumerate(cabs, 1):
        cel = ws.cell(1, c, t)
        cel.font = _f(bold=True, color="FFFFFF")
        cel.fill = _CAB_FILL
    ws.freeze_panes = "A2"
    # Agrupación nativa de Excel (botones +/− al margen): el encabezado de cada
    # categoría queda ARRIBA de su bloque, así que el resumen va arriba.
    ws.sheet_properties.outlinePr.summaryBelow = False
    fila = 2
    # El archivo abre TOTALMENTE plegado: solo las categorías principales
    # (Eduardo, 05-ago); todo lo demás se abre con los +. Tope Excel: 7 niveles.
    _VISIBLE_HASTA = 0  # profundidad máxima visible al abrir

    def _outline(r: int, nivel: int, oculta: bool) -> None:
        rd = ws.row_dimensions[r]
        rd.outlineLevel = min(nivel, 7)
        if oculta:
            rd.hidden = True

    def pinta(n: _Nodo, ruta: str, depth: int) -> None:
        nonlocal fila
        r0 = fila
        a = acum[ruta]
        cel = ws.cell(r0, 1, f"{n.nombre} ({a['pubs']} pub)")
        cel.font = _f(bold=True)
        cel.alignment = Alignment(indent=depth)
        fill = PatternFill("solid", fgColor=_NIVEL_FILL[min(depth, 3)])
        for c in range(1, _N + 1):
            ws.cell(r0, c).fill = fill
        if depth:  # las raíces (nivel 0) no pertenecen a ningún grupo
            _outline(r0, depth, depth > _VISIBLE_HASTA)
        if depth >= _VISIBLE_HASTA:  # sus hijos abren plegados
            ws.row_dimensions[r0].collapsed = True
        fila += 1
        items = sorted((p for h in n.hojas
                        for p in pubs_por_cat.get(str(h["category_id"]), [])),
                       key=lambda p: -float(p["venta"] or 0))
        for p in items:
            ws.cell(fila, 1, (p.get("sku") or "(sin SKU)")).font = _f()
            ws.cell(fila, 1).alignment = Alignment(indent=depth + 1)
            ws.cell(fila, 2, TIENDA.get(p.get("cuenta") or "", p.get("cuenta"))).font = _f()
            ws.cell(fila, 3, p.get("titulo") or "").font = _f()
            ws.cell(fila, 4, p.get("item_id") or "").font = _f()
            ws.cell(fila, 5, p.get("situacion") or "").font = _f()
            ws.cell(fila, 6, int(p.get("uds") or 0)).font = _f()
            ws.cell(fila, 6).number_format = _INT
            ws.cell(fila, 7, float(p.get("venta") or 0)).font = _f()
            ws.cell(fila, 7).number_format = _MONEY
            # Costos: en blanco cuando el producto no tiene costo capturado.
            # Sin Ganancia ni Margen %: ver el encabezado del módulo.
            # `costo_final_real` lo arma el router con el envío ya resuelto
            # (real de ML donde lo hay). Si no vino, se cae al de costing.
            cf = p.get("costo_final_real")
            cf = _costo_final(p) if cf is None else float(cf)
            if cf is not None:
                ws.cell(fila, 8, float(p["costo_base"])).font = _f()
                ws.cell(fila, 9, float(p.get("comision") or 0)).font = _f()
                if p.get("envio") is not None:
                    ws.cell(fila, 10, float(p["envio"])).font = _f()
                ws.cell(fila, 11, round(cf, 2)).font = _f()
                for c in (8, 9, 10, 11):
                    ws.cell(fila, c).number_format = _MONEY
            if p.get("precio") is not None:
                ws.cell(fila, 12, float(p["precio"])).font = _f()
                ws.cell(fila, 12).number_format = _MONEY
            ws.cell(fila, 13, p.get("primera_venta") or "").font = _f()
            ws.cell(fila, 14, p.get("ultima_venta") or "").font = _f()
            org = p.get("envio_origen") or ""
            cel_org = ws.cell(fila, 15, org)
            cel_org.font = _f(size=9, bold=(org == "ML real"))
            if org and org != "ML real":
                cel_org.fill = _AVISO_FILL
            aviso = diagnosticar(p, p.get("venta"), p.get("costo_base"))
            if aviso:
                ws.cell(fila, 16, aviso).font = _f(size=9)
                for c in (8, 9, 10, 11):
                    if ws.cell(fila, c).value is None:
                        ws.cell(fila, c).fill = _AVISO_FILL
            # las publicaciones cuelgan un nivel bajo su categoría
            _outline(fila, depth + 1, depth + 1 > _VISIBLE_HASTA)
            fila += 1
        for h in sorted(n.hijos.values(),
                        key=lambda x: -acum[f"{ruta}›{x.nombre}"]["venta"]):
            pinta(h, f"{ruta}›{h.nombre}", depth + 1)
        r1 = fila - 1
        if r1 > r0:
            # SUBTOTAL(9,…) ignora las filas plegadas y los subtotales anidados,
            # así que cada nivel suma bien sin contar dos veces.
            for col, fmt in (("F", _INT), ("G", _MONEY), ("H", _MONEY),
                             ("I", _MONEY), ("J", _MONEY), ("K", _MONEY)):
                c = ord(col) - 64
                ws.cell(r0, c, f"=SUBTOTAL(9,{col}{r0 + 1}:{col}{r1})").font = _f(bold=True)
                ws.cell(r0, c).number_format = fmt
        else:  # nodo sin publicaciones desglosadas: valores del acumulado
            ws.cell(r0, 6, a["uds"]).font = _f(bold=True)
            ws.cell(r0, 6).number_format = _INT
            ws.cell(r0, 7, round(a["venta"], 2)).font = _f(bold=True)
            ws.cell(r0, 7).number_format = _MONEY

    for r in orden_raiz:
        pinta(raices[r], r, 0)
    for c, w in {1: 42, 2: 9, 3: 52, 4: 15, 5: 10, 6: 10, 7: 12, 8: 12,
                 9: 12, 10: 11, 11: 12, 12: 10, 13: 11, 14: 11, 15: 12,
                 16: 80}.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ult = fila - 1

    # ── hoja Resumen ─────────────────────────────────────────────────────────
    rs = wb["Sheet"]
    rs.title = "Resumen"
    rs["A1"] = "Ventas y costos por categoría"
    rs["A1"].font = _f(bold=True, size=12)
    rs["B1"] = f"{desde} → {hasta}"
    rs["B1"].font = _f(bold=True)

    # RANGO PEDIDO ≠ RANGO CON DATOS. La captura de pedidos arrancó de verdad a
    # finales de junio de 2026: antes de eso `channel.orders` tiene 186 filas en
    # total (verificado el 7-ago contra `pedidos_ml`, el registro viejo, que
    # coincide — no falta historia, no existe). Sin este aviso, pedir "Histórico
    # (400 días)" devuelve un libro que parece cubrir un año y cubre siete
    # semanas, y los meses vacíos se leen como meses malos.
    fechas = sorted(str(v["fecha"]) for v in ventas if v.get("fecha"))
    if not fechas:
        rs["A2"] = (f"SIN VENTAS en el rango pedido ({desde} → {hasta}). "
                    f"El libro va vacío: no es que no haya margen, es que no "
                    f"hay pedidos capturados en esas fechas.")
        rs["A2"].font = _f(bold=True, size=9)
        rs["A2"].fill = _AVISO_FILL
    elif fechas[0] > desde:
        dias_reales = len(set(fechas))
        rs["A2"] = (f"OJO CON EL RANGO: se pidió desde {desde}, pero la primera "
                    f"venta capturada es del {fechas[0]} (hay ventas en "
                    f"{dias_reales} días distintos, hasta {fechas[-1]}). "
                    f"Los meses anteriores no salen bajos: salen sin captura. "
                    f"Compara solo dentro del rango con datos.")
        rs["A2"].font = _f(bold=True, size=9)
        rs["A2"].fill = _AVISO_FILL
    rs["C1"] = f"Cuenta: {TIENDA.get(cuenta or '', cuenta) or 'todas'}"
    rs["C1"].font = _f(bold=True)
    cob = ""
    if censo_envio:
        n = sum(censo_envio.values())
        if n:
            cob = (f" ENVÍO: {censo_envio['reales']:,} de {n:,} renglones "
                   f"({censo_envio['reales'] / n * 100:.0f}%) traen el cobro "
                   f"REAL que hizo ML por el embarque; el resto va con el "
                   f"estimado de costing. La columna 'Origen envío' lo dice "
                   f"renglón por renglón.")
    rs["D1"] = ("Todo el libro sale de los PEDIDOS del período. Costo base = "
                "producto + flete de importación; costo final le suma la "
                "comisión REAL de Mercado Libre y el envío." + cob + " Este "
                "reporte NO calcula ganancia ni margen: la base de costos tiene "
                "defectos medidos (precios placeholder, pesos de caja como "
                "pieza) que harían ver un margen como un hecho sin serlo. "
                "Compara Venta con costo contra Ventas $ para saber sobre qué "
                "parte de la categoría estás mirando costos.")
    rs["D1"].font = _f(italic=True, size=9)

    cabs_r = ["Categoría Principal", "SKUs con venta", "Ventas Uds", "Ventas $",
              "% Ventas $", "Costo base", "Costo final", "Venta con costo",
              "Publicaciones", "Activas"]
    for c, t in enumerate(cabs_r, 1):
        cel = rs.cell(3, c, t)
        cel.font = _f(bold=True, color="FFFFFF")
        cel.fill = _CAB_FILL
    rs.freeze_panes = "A4"
    tot_row = 4 + len(orden_raiz)
    for i, r in enumerate(orden_raiz):
        rr = 4 + i
        a = acum[r]
        rs.cell(rr, 1, r).font = _f()
        rs.cell(rr, 2, a["skus"]).font = _f()
        rs.cell(rr, 3, a["uds"]).font = _f()
        rs.cell(rr, 4, round(a["venta"], 2)).font = _f()
        rs.cell(rr, 5, f"=IF($D${tot_row}=0,0,D{rr}/$D${tot_row})").font = _f()
        rs.cell(rr, 6, round(a["costo_base"], 2)).font = _f()
        rs.cell(rr, 7, round(a["costo_final"], 2)).font = _f()
        rs.cell(rr, 8, round(a["venta_con_costo"], 2)).font = _f()
        rs.cell(rr, 9, a["pubs"]).font = _f()
        rs.cell(rr, 10, a["activas"]).font = _f()
        for c, fmt in ((2, _INT), (3, _INT), (4, _MONEY), (5, _PCT),
                       (6, _MONEY), (7, _MONEY), (8, _MONEY), (9, _INT),
                       (10, _INT)):
            rs.cell(rr, c).number_format = fmt
    rs.cell(tot_row, 1, "TOTAL").font = _f(bold=True)
    for col, fmt in (("B", _INT), ("C", _INT), ("D", _MONEY), ("F", _MONEY),
                     ("G", _MONEY), ("H", _MONEY), ("I", _INT), ("J", _INT)):
        rs.cell(tot_row, ord(col) - 64,
                f"=SUM({col}4:{col}{tot_row - 1})").font = _f(bold=True)
        rs.cell(tot_row, ord(col) - 64).number_format = fmt
    rs.cell(tot_row, 5, f"=IF($D${tot_row}=0,0,1)").font = _f(bold=True)
    rs.cell(tot_row, 5).number_format = _PCT
    rs.cell(3, 2).comment = Comment(
        "SKUs distintos con venta en el período. En categorías con "
        "sub-clasificación doble un SKU puede contar en dos ramas (misma "
        "convención que la página).", "OMNICANAL")
    rs.cell(3, 8).comment = Comment(
        "Parte de las Ventas $ cuyo producto SÍ tiene costo capturado. Es el "
        "medidor de COBERTURA: si queda muy por debajo de Ventas $, las "
        "columnas de costo de esta fila describen una muestra, no la "
        "categoría completa.", "OMNICANAL")
    for c, w in {1: 32, 2: 13, 3: 11, 4: 13, 5: 10, 6: 13, 7: 13, 8: 14,
                 9: 13, 10: 9}.items():
        rs.column_dimensions[get_column_letter(c)].width = w

    # ── hoja Ventas ──────────────────────────────────────────────────────────
    # Una fila por LÍNEA vendida: el grano más fino del libro. Sustituye al CSV
    # de márgenes que se descargaba aparte. Sin fórmulas ni agrupaciones: está
    # pensada para tabla dinámica y para auditar una cifra que no cuadre.
    vs = wb.create_sheet("Ventas")
    cabs_v = ["Fecha", "Canal", "Cuenta", "Pedido", "SKU", "Título", "Cant.",
              "Precio unit.", "Ingreso", "Comisión ML", "Envío",
              "Costo base", "Costo final", "FULL", "Estado", "Origen envío",
              "Diagnóstico"]
    for c, t in enumerate(cabs_v, 1):
        cel = vs.cell(1, c, t)
        cel.font = _f(bold=True, color="FFFFFF")
        cel.fill = _CAB_FILL
    vs.freeze_panes = "A2"
    vs.auto_filter.ref = f"A1:{get_column_letter(len(cabs_v))}1"
    for i, v in enumerate(ventas):
        r = 2 + i
        ingreso = float(v.get("ingreso") or 0)
        cfin = v.get("costo_final")
        vs.cell(r, 1, v.get("fecha") or "").font = _f()
        vs.cell(r, 2, v.get("canal") or "").font = _f()
        vs.cell(r, 3, TIENDA.get(v.get("cuenta") or "", v.get("cuenta"))).font = _f()
        vs.cell(r, 4, str(v.get("pedido") or "")).font = _f()
        vs.cell(r, 5, v.get("sku") or "").font = _f()
        vs.cell(r, 6, v.get("titulo") or "").font = _f()
        vs.cell(r, 7, int(v.get("cantidad") or 0)).font = _f()
        vs.cell(r, 8, float(v.get("precio_unitario") or 0)).font = _f()
        vs.cell(r, 9, ingreso).font = _f()
        vs.cell(r, 10, float(v.get("comision_ml") or 0)).font = _f()
        # El envío es el COBRO REAL de ML cuando lo tenemos, y el estimado de
        # costing si no; la columna "Origen envío" dice cuál de los dos, porque
        # una columna que mezcla fuentes en silencio es la misma trampa que el
        # margen que se retiró. Va VACÍO cuando no hay ninguno de los dos:
        # antes iba `or 0` y eso volvía indistinguible "el envío costó cero"
        # de "no lo sabemos".
        if v.get("envio") is not None:
            vs.cell(r, 11, float(v["envio"])).font = _f()
        if v.get("costo_base") is not None:
            vs.cell(r, 12, float(v["costo_base"])).font = _f()
        if cfin is not None:
            vs.cell(r, 13, float(cfin)).font = _f()
        vs.cell(r, 14, "sí" if v.get("full") else "no").font = _f()
        vs.cell(r, 15, v.get("estado") or "").font = _f()
        org = v.get("envio_origen") or ""
        cel_org = vs.cell(r, 16, org)
        cel_org.font = _f(size=9, bold=(org == "ML real"))
        if org != "ML real":
            cel_org.fill = _AVISO_FILL
        aviso = diagnosticar(v, ingreso, v.get("costo_base"))
        if aviso:
            vs.cell(r, 17, aviso).font = _f(size=9)
            for c in (11, 12, 13):
                if vs.cell(r, c).value is None:
                    vs.cell(r, c).fill = _AVISO_FILL
        for c, fmt in ((7, _INT), (8, _MONEY), (9, _MONEY), (10, _MONEY),
                       (11, _MONEY), (12, _MONEY), (13, _MONEY)):
            vs.cell(r, c).number_format = fmt
    for c, w in {1: 11, 2: 14, 3: 9, 4: 16, 5: 18, 6: 52, 7: 7, 8: 12, 9: 12,
                 10: 12, 11: 11, 12: 12, 13: 12, 14: 6, 15: 12, 16: 12,
                 17: 80}.items():
        vs.column_dimensions[get_column_letter(c)].width = w

    _ = ult  # (referencia futura: filas totales de la hoja Categorias)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
