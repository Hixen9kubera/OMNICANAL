"""
reporte_inventario_xlsx.py — Inventario ACCIONABLE, no un volcado del almacén.

DOS HOJAS, dos problemas opuestos (Eduardo, 7-ago):

  INMOVILIZADO — hay stock en FULL y no vende. Ahí no es solo capital
    detenido: paga renta a Mercado Libre todos los días. El mercado no lo
    quiere. Censo del 7-ago: 14,873 unidades en FULL sin una sola venta en 30
    días —el 39%% de todo lo que tenemos allá— y las mayores NUNCA han vendido
    una pieza (JUGU-0261-LIL: 272 en FULL + 648 en bodega, cero ventas
    históricas).

  INVISIBLE — vende, pero ninguna publicación está activa. El mercado sí lo
    quiere y no se lo estamos ofreciendo. Caso canónico TEC-0393-ROS: 291
    unidades vendidas en 30 días, 2,394 en bodega, las dos publicaciones
    pausadas.

Es el problema más barato de arreglar de los dos: no hay que comprar, mover ni
liquidar nada — solo reactivar.

SIN VALORIZAR EN DINERO, a propósito. Es la misma trampa del margen que se
retiró del otro reporte: `costo_producto` es un precio USD×19 de relleno en
~30%% del catálogo, así que un total de "inventario valuado en $X" sería
ficción con formato de hecho. Lo que sí se puede medir con datos confiables es
CUÁNTO hay, DÓNDE está y CUÁNTO TIEMPO lleva sin moverse.
"""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_CAB_FILL = PatternFill("solid", fgColor="1F3864")
_AVISO_FILL = PatternFill("solid", fgColor="FFF2CC")
_GRAVE_FILL = PatternFill("solid", fgColor="FCE4E4")
_INT = "#,##0"
_MXN = '"$"#,##0'

# WooCommerce NO va aquí a propósito: no es un canal de venta sino nuestro
# puente de registro, y la consulta ya lo excluye de la lista de cuentas. Si
# alguna vez se colara, saldría como "GENERAL" en mayúsculas — una señal
# visible, mejor que pintarlo como si fuera una tienda más.
TIENDA = {"BEKURA": "Bekura", "SANCORFASHION": "Sancor", "AMAZON": "Amazon"}


def _f(**kw) -> Font:
    return Font(name="Arial", **{"size": 10, **kw})


def _cuentas(v: Any) -> str:
    if not v:
        return ""
    return ", ".join(TIENDA.get(c, c) for c in v if c)


def _donde(f: dict) -> str:
    """En qué VARIANTE y en qué cuenta están de verdad las piezas de FULL.

    El renglón se nombra con el SKU padre, y el padre de un producto con
    variantes no vende nunca: vende la variante. Sin esta columna, revisar el
    reporte llevaba a buscar las ventas del padre, no encontrar ninguna y
    concluir que el dato estaba mal (Eduardo, 14-ago, sobre CAM-0030).
    """
    det = f.get("full_detalle") or []
    if not det:
        return ""
    # La columna solo gana su lugar cuando el stock vive en un SKU DISTINTO del
    # que nombra el renglón. Si todas las piezas son del mismo código, esto
    # repetiría la columna 1 y el reparto por cuenta que ya dan «FULL Bekura» y
    # «FULL Sancor» — ruido, no información.
    if all(d.get("sku") == f.get("sku") for d in det):
        return ""
    return "   |   ".join(
        f"{d.get('sku')} · {TIENDA.get(d.get('cuenta'), d.get('cuenta'))}"
        f" · {int(d.get('uds') or 0):,}" for d in det)


def _encabezar(ws, cabs: list[str], nota: str, anchos: dict[int, int]) -> None:
    """Título explicativo en la fila 1 y encabezados en la 3.

    La nota va DENTRO del archivo porque el Excel se comparte fuera del panel:
    un criterio que solo vive en la pantalla no viaja con el adjunto.
    """
    ws["A1"] = nota
    ws["A1"].font = _f(italic=True, size=9)
    ws["A1"].alignment = Alignment(vertical="top")
    for c, t in enumerate(cabs, 1):
        cel = ws.cell(3, c, t)
        cel.font = _f(bold=True, color="FFFFFF")
        cel.fill = _CAB_FILL
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(cabs))}3"
    for c, w in anchos.items():
        ws.column_dimensions[get_column_letter(c)].width = w


def _mxn(filas: list[dict]) -> float:
    """Suma del valor a precio de venta de un conjunto de renglones."""
    return sum(float(f.get("valor_full") or 0) for f in filas)


def _celda_valor(ws, r: int, col: int, f: dict) -> None:
    """El valor del renglón, y la advertencia cuando todavía es precio de LISTA.

    `precio_crudo` significa que a NINGUNA de las publicaciones de ese SKU se le
    ha leído el precio con promoción, así que la cifra corre alto —medido, 1.71x
    lo que de verdad se transa—. Se muestra igual, en ámbar: esconderla dejaría
    el total sin cuadrar, y el reporte tiene que ser auditable como el de la CAM.
    """
    ws.cell(r, col, round(float(f.get("valor_full") or 0), 2)).font = _f()
    ws.cell(r, col).number_format = _MXN
    if f.get("precio_crudo"):
        ws.cell(r, col).fill = _AVISO_FILL


def _hoja_inmovilizado(wb: Workbook, filas: list[dict], dias: int) -> None:
    ws = wb.create_sheet("Inmovilizado")
    cabs = ["SKU", "Título", "Dónde está el stock", "En FULL", "FULL Bekura",
            "FULL Sancor", "En bodega propia", "Publicaciones activas",
            "Pausadas", "Cuentas", "Última venta", "Días sin vender",
            "Valor en FULL", "Diagnóstico"]
    total_uds = sum(int(f.get("full_total") or 0) for f in filas)
    nunca = sum(1 for f in filas if not f.get("ultima_venta"))
    _encabezar(ws, cabs,
               f"INMOVILIZADO — stock en FULL que NO vendió una sola pieza en "
               f"los últimos {dias} días. Ahí paga almacenaje a Mercado Libre "
               f"todos los días, venda o no. {len(filas):,} SKUs, "
               f"{total_uds:,} unidades; {nunca:,} nunca han vendido nada. "
               f"Ordenado por unidades: arriba está lo que más renta paga sin "
               f"devolver nada. Vale {_mxn(filas):,.0f} pesos A PRECIO DE "
               f"VENTA: no al costo —que no es de fiar en ~30% del catálogo— "
               f"sino a lo que se cobraría si se vendiera hoy, con la promoción "
               f"de Mercado Libre ya aplicada. No incluye el stock parado en "
               f"bodega propia (ese no paga renta). En un producto con variantes "
               f"el renglón es la FAMILIA COMPLETA —el SKU de la izquierda es "
               f"el padre, que nunca vende por sí mismo— y la columna «Dónde "
               f"está el stock» dice en qué variante y en qué cuenta están las "
               f"piezas.",
               {1: 22, 2: 42, 3: 46, 4: 10, 5: 12, 6: 12, 7: 16, 8: 18,
                9: 10, 10: 16, 11: 12, 12: 14, 13: 15, 14: 62})
    for i, f in enumerate(filas):
        r = 4 + i
        dias_sin = f.get("dias_sin_vender")
        ws.cell(r, 1, f.get("sku") or "").font = _f()
        ws.cell(r, 2, (f.get("titulo") or "")[:120]).font = _f()
        ws.cell(r, 3, _donde(f)).font = _f(size=9)
        for c, k in ((4, "full_total"), (5, "full_bk"), (6, "full_sc"),
                     (7, "propio"), (8, "activas"), (9, "pausadas")):
            ws.cell(r, c, int(f.get(k) or 0)).font = _f()
            ws.cell(r, c).number_format = _INT
        ws.cell(r, 10, _cuentas(f.get("cuentas"))).font = _f()
        ws.cell(r, 11, f.get("ultima_venta") or "nunca").font = _f()
        if dias_sin is not None:
            ws.cell(r, 12, int(dias_sin)).font = _f()
            ws.cell(r, 12).number_format = _INT
        # Nunca vendido es distinto de "dejó de venderse": no es que se
        # enfriara, es que la compra no tenía mercado. Se pinta distinto.
        #
        # El diagnóstico DESCRIBE, no receta (Eduardo, 7-ago): qué hacer con un
        # inmovilizado —sacarlo de FULL, liquidarlo, dejar de comprarlo— depende
        # de temporada, contrato y planes que el reporte no conoce. Aquí solo va
        # el hecho. En Invisible sí se sugiere, porque ahí la acción es una sola
        # y no admite matices: la publicación está apagada teniendo stock.
        if not f.get("ultima_venta"):
            aviso = ("NUNCA HA VENDIDO — no dejó de venderse: jamás vendió una "
                     "pieza, y aun así ocupa lugar en FULL")
            ws.cell(r, 11).fill = _GRAVE_FILL
        else:
            aviso = (f"SIN VENTA EN {dias} DÍAS — lleva {int(dias_sin):,} días "
                     f"desde su última venta y sigue ocupando FULL")
            ws.cell(r, 11).fill = _AVISO_FILL
        # En una familia, "nunca vendió" es de la familia ENTERA: ni el padre ni
        # una sola variante. Decirlo evita la lectura de que solo se revisó el
        # padre — que es justo lo que el SKU de la izquierda sugiere.
        if int(f.get("variantes") or 0):
            aviso += (f" · la cuenta cubre al padre y a sus "
                      f"{int(f['variantes'])} variantes")
        _celda_valor(ws, r, 13, f)
        if f.get("precio_crudo"):
            aviso += " · VALUADO A PRECIO DE LISTA: la cifra corre alto"
        ws.cell(r, 14, aviso).font = _f(size=9)


def _hoja_invisible(wb: Workbook, filas: list[dict], dias: int) -> None:
    ws = wb.create_sheet("Invisible")
    cabs = ["SKU", "Título", f"Vendió ({dias}d)", "Uds/día", "En bodega propia",
            "En FULL", "En FBA", "Stock total", "Cobertura (días)",
            "Publicaciones pausadas", "Cuentas", "Última venta",
            "Valor en FULL", "Diagnóstico"]
    uds = sum(int(f.get("uds_periodo") or 0) for f in filas)
    _encabezar(ws, cabs,
               f"INVISIBLE — vendió en los últimos {dias} días y hoy NO tiene "
               f"una sola publicación activa, teniendo stock con qué surtir. "
               f"Demanda probada con el aparador cerrado. {len(filas):,} SKUs, "
               f"{uds:,} unidades vendidas que hoy no se pueden repetir. "
               f"Lo pausado SIN stock queda FUERA a propósito: está agotado, "
               f"que es la razón correcta para pausar, y pertenece a Reponer. "
               f"Es el problema más barato de arreglar: no hay que comprar ni "
               f"mover nada, solo reactivar. Lo que está en FULL vale "
               f"{_mxn(filas):,.0f} pesos a precio de venta y hoy no se le "
               f"ofrece a nadie.",
               {1: 22, 2: 46, 3: 12, 4: 9, 5: 16, 6: 10, 7: 9, 8: 12, 9: 15,
                10: 20, 11: 16, 12: 12, 13: 15, 14: 64})
    for i, f in enumerate(filas):
        r = 4 + i
        u = int(f.get("uds_periodo") or 0)
        stock = int(f.get("stock_total") or 0)
        vel = u / dias if dias else 0
        ws.cell(r, 1, f.get("sku") or "").font = _f()
        ws.cell(r, 2, (f.get("titulo") or "")[:120]).font = _f()
        ws.cell(r, 3, u).font = _f(bold=True)
        ws.cell(r, 4, round(vel, 2)).font = _f()
        ws.cell(r, 4).number_format = "0.00"
        for c, k in ((5, "propio"), (6, "full_total"), (7, "fba")):
            ws.cell(r, c, int(f.get(k) or 0)).font = _f()
        ws.cell(r, 8, stock).font = _f()
        # Cobertura = cuánto duraría el stock al ritmo al que SE VENDÍA. Solo
        # tiene sentido aquí porque estas filas sí tienen ventas; en
        # Inmovilizado la división sería entre cero.
        if vel > 0:
            ws.cell(r, 9, round(stock / vel)).font = _f()
            ws.cell(r, 9).number_format = _INT
        ws.cell(r, 10, int(f.get("pausadas") or 0)).font = _f()
        ws.cell(r, 11, _cuentas(f.get("cuentas"))).font = _f()
        ws.cell(r, 12, f.get("ultima_venta") or "").font = _f()
        for c in (3, 5, 6, 7, 8, 10):
            ws.cell(r, c).number_format = _INT
        _celda_valor(ws, r, 13, f)
        ws.cell(r, 14,
                f"PAUSADA CON STOCK — vendió {u:,} unidades en {dias} días y "
                f"hoy tiene {stock:,} disponibles sin ninguna publicación "
                f"activa; reactivar o entender por qué se pausó"
                + (" · su valor está a precio de LISTA: corre alto"
                   if f.get("precio_crudo") else "")
                ).font = _f(size=9)
        ws.cell(r, 8).fill = _AVISO_FILL


def _hoja_valor(wb: Workbook, filas: list[dict], dias: int) -> None:
    """TODO el inventario en FULL valuado a precio de venta.

    Replica el corte que la CAM armaba a mano cruzando el reporte
    `stock_general_full` de ML con los precios de la tienda pública. Al comparar
    su archivo del 13-ago contra el nuestro reconstruido a esa misma hora, las
    UNIDADES coincidieron al 98.5% (298 de 357 publicaciones con el número
    idéntico); lo que no coincidía era el precio, y el error era nuestro.

    Las otras dos hojas son subconjuntos de esta: Inmovilizado es lo que además
    no vendió, Invisible lo que además no tiene publicación activa.
    """
    ws = wb.create_sheet("Valor en FULL")
    cabs = ["SKU", "Título", "Dónde está el stock", "En FULL", "FULL Bekura",
            "FULL Sancor", "Valor a precio de venta", "Activas", "Pausadas",
            "Cuentas", f"Vendidas ({dias}d)", "Última venta", "Aviso"]
    val = _mxn(filas)
    uds = sum(int(f.get("full_total") or 0) for f in filas)
    crudas = [f for f in filas if f.get("precio_crudo")]
    sin_precio = sum(int(f.get("uds_sin_precio") or 0) for f in filas)
    _encabezar(ws, cabs,
               f"VALOR DEL INVENTARIO EN FULL a precio de venta — lo que se "
               f"cobraría si se vendiera hoy, con la promoción de Mercado Libre "
               f"ya aplicada. {len(filas):,} publicaciones con stock, "
               f"{uds:,} unidades, {val:,.0f} pesos. NO es el costo: el costo "
               f"capturado no es de fiar en cerca de un tercio del catálogo, y "
               f"por eso este libro nunca tuvo pesos. El precio de anaquel no "
               f"depende de ese dato. "
               + (f"OJO: {len(crudas):,} publicaciones "
                  f"({sum(int(f.get('full_total') or 0) for f in crudas):,} "
                  f"unidades, {_mxn(crudas):,.0f} pesos) siguen valuadas con "
                  f"precio de LISTA porque todavía no se les lee el precio con "
                  f"promoción: van en ámbar y su valor corre ALTO —mediana "
                  f"medida, 1.71 veces lo que de verdad se transa—. El total "
                  f"baja cuando el sync termine de observarlas. "
                  if crudas else
                  "Todas las publicaciones tienen leído su precio con "
                  "promoción. ")
               + (f"{sin_precio:,} unidades no tienen ningún precio y quedan "
                  f"FUERA del total." if sin_precio else ""),
               {1: 22, 2: 40, 3: 46, 4: 10, 5: 12, 6: 12, 7: 22, 8: 10,
                9: 10, 10: 16, 11: 14, 12: 12, 13: 60})
    for i, f in enumerate(filas):
        r = 4 + i
        ws.cell(r, 1, f.get("sku") or "").font = _f()
        ws.cell(r, 2, (f.get("titulo") or "")[:120]).font = _f()
        ws.cell(r, 3, _donde(f)).font = _f(size=9)
        for c, k in ((4, "full_total"), (5, "full_bk"), (6, "full_sc"),
                     (8, "activas"), (9, "pausadas"), (11, "uds_periodo")):
            ws.cell(r, c, int(f.get(k) or 0)).font = _f()
            ws.cell(r, c).number_format = _INT
        _celda_valor(ws, r, 7, f)
        ws.cell(r, 10, _cuentas(f.get("cuentas"))).font = _f()
        ws.cell(r, 12, f.get("ultima_venta") or "nunca").font = _f()
        avisos = []
        # PRIMERO qué es el renglón. Un padre no vende NUNCA —vende su
        # variante— así que sin decirlo, "0 vendidas · nunca" se lee como un
        # error del reporte en vez de como el hecho de la familia entera.
        # Medido el 20-ago-2026 sobre todo el histórico: los SKUs padre suman
        # 25 unidades vendidas contra 2,952 de los hijos.
        variantes = int(f.get("variantes") or 0)
        if variantes:
            avisos.append(f"FAMILIA de {variantes} variantes — el SKU de la "
                          f"izquierda es el PADRE y no vende por sí mismo; las "
                          f"piezas y las ventas son de sus variantes (columna "
                          f"«Dónde está el stock»). Los números de este "
                          f"renglón son de la familia COMPLETA")
        if f.get("precio_crudo"):
            avisos.append("VALUADO A PRECIO DE LISTA — todavía no se lee su "
                          "precio con promoción; la cifra corre alto")
        if int(f.get("uds_sin_precio") or 0):
            avisos.append(f"{int(f['uds_sin_precio']):,} unidades sin ningún "
                          f"precio: NO están en el valor de la izquierda")
        if not int(f.get("uds_periodo") or 0):
            avisos.append(f"sin vender en {dias} días"
                          + (" —ni el padre ni una sola variante—" if variantes else "")
                          + " · también sale en Inmovilizado")
        ws.cell(r, 13, " · ".join(avisos)).font = _f(size=9)


def construir(inmovilizado: list[dict], invisible: list[dict],
              valor: list[dict], dias: int, cuenta: str | None) -> bytes:
    wb = Workbook()
    portada = wb["Sheet"]
    portada.title = "Cómo leer"
    portada["A1"] = "Inventario en FULL: valor y acciones"
    portada["A1"].font = _f(bold=True, size=12)
    portada["A2"] = (f"Período: últimos {dias} días · "
                     f"Cuenta: {TIENDA.get(cuenta or '', cuenta) or 'todas'}")
    portada["A2"].font = _f(bold=True)
    lineas = [
        "",
        "Este libro NO es el inventario completo: son las dos poblaciones "
        "sobre las que se puede actuar hoy, y son problemas opuestos.",
        "",
        f"INMOVILIZADO ({len(inmovilizado):,} SKUs) — el mercado no lo quiere. "
        "Hay stock en FULL y no vende, así que paga renta todos los días sin "
        "devolver nada. La acción es sacarlo de FULL, liquidarlo o dejar de "
        "comprarlo.",
        "",
        f"INVISIBLE ({len(invisible):,} SKUs) — el mercado sí lo quiere y no se "
        "lo estamos ofreciendo. Vendió, tiene stock, y ninguna publicación "
        "está activa. La acción es reactivar, o entender por qué se pausó.",
        "",
        f"VALOR EN FULL ({len(valor):,} publicaciones) — todo lo que ocupa "
        f"FULL, venda o no, valuado en {_mxn(valor):,.0f} pesos. Las otras dos "
        f"hojas son subconjuntos de esta.",
        "",
        "EL VALOR VA A PRECIO DE VENTA, NO A COSTO. Este libro no tenía pesos "
        "porque el costo capturado es un precio de lista en dólares en cerca "
        "de un tercio del catálogo, y valorizar así daría una cifra inventada. "
        "El precio de anaquel no depende de ese dato: es lo que se cobraría si "
        "se vendiera hoy, con la promoción de Mercado Libre ya aplicada. Lo "
        "que todavía se valuó con precio de LISTA va en ámbar y corre alto "
        "—mediana medida, 1.71 veces lo que de verdad se transa—.",
        "",
        "El stock propio se cuenta UNA vez por SKU, no por publicación: la "
        "misma bodega se ve desde cada publicación, y sumarlas contaría la "
        "misma pieza varias veces. FULL y FBA sí son por cuenta.",
    ]
    for i, t in enumerate(lineas, start=4):
        portada.cell(i, 1, t).font = _f(size=10)
        portada.cell(i, 1).alignment = Alignment(wrap_text=True, vertical="top")
    portada.column_dimensions["A"].width = 118

    _hoja_valor(wb, valor, dias)
    _hoja_inmovilizado(wb, inmovilizado, dias)
    _hoja_invisible(wb, invisible, dias)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
